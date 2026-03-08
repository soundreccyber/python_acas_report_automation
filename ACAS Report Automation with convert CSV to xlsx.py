print("SCRIPT STARTED")

import os
import requests
import csv
import sys
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# ==============================
# CONFIG
# ==============================

SECURITY_CENTER = "https://security_Center_URL/rest"
USERNAME = "Your ID"
PASSWORD = "Your_Password"

EXPORT_DIR = r"C:\Your_Export_Folder_Location"
LOG_FILE = r"C:\Your_Log_Folder_Location\Your_log_file.log"

# Existing report result names to download
EXISTING_REPORTS = [
    {"name": "Vulnerability CSV Report", "ext": "csv"},
    {"name": "Critical and Exploitable Vulnerabilities Report", "ext": "pdf"},
    {"name": "Monthly Executive Report", "ext": "pdf"},
    {"name": "Remediation Instructions by Host Report", "ext": "pdf"},
]

requests.packages.urllib3.disable_warnings()

# ==============================
# LOGGING
# ==============================

def write_log(message):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    entry = f"{timestamp} {message}"
    print(entry)

    os.makedirs(os.path.dirname(LOG_FILE), exist_ok=True)
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(entry + "\n")

# ==============================
# OUTPUT DIRECTORY
# ==============================

def get_output_dir():
    now = datetime.now()
    year_folder = now.strftime("%Y")
    month_folder = now.strftime("%m")

    output_dir = os.path.join(EXPORT_DIR, year_folder, month_folder)
    os.makedirs(output_dir, exist_ok=True)

    return output_dir

# ==============================
# LOGIN
# ==============================

def get_token(session):
    write_log("Requesting SecurityCenter token...")

    url = f"{SECURITY_CENTER}/token"
    payload = {
        "username": USERNAME,
        "password": PASSWORD
    }

    r = session.post(url, json=payload, verify=False)
    data = r.json()

    print("LOGIN RESPONSE:", r.text)

    if data.get("error_code") != 0:
        raise RuntimeError(f"Login failed: {data}")

    token = str(data["response"]["token"])
    write_log(f"Token received: {token}")
    return token

# ==============================
# OPTIONAL SCAN CHECK
# ==============================

def get_latest_completed_scan(session, headers):
    write_log("Retrieving latest completed scan...")

    url = f"{SECURITY_CENTER}/scan"
    r = session.get(url, headers=headers, verify=False)
    data = r.json()

    if data.get("error_code") != 0:
        raise RuntimeError(f"Failed to retrieve scans: {data}")

    scans = data["response"].get("usable", [])
    completed = [s for s in scans if str(s.get("status")) in ("Completed", "0")]

    if not completed:
        raise RuntimeError("No completed scans found")

    latest = sorted(completed, key=lambda x: int(x["id"]), reverse=True)[0]
    write_log(f"Latest completed scan: ID={latest['id']} Name={latest['name']}")
    return latest

# ==============================
# EXISTING REPORT RESULT LOOKUP
# ==============================

def get_existing_reports_by_name(session, headers, report_name):
    write_log(f"Looking up existing report results for: {report_name}")

    url = f"{SECURITY_CENTER}/report"
    r = session.get(url, headers=headers, verify=False)
    data = r.json()

    if data.get("error_code") != 0:
        raise RuntimeError(f"Failed to retrieve reports: {data}")

    reports = data["response"].get("usable", [])
    matches = [rpt for rpt in reports if rpt.get("name") == report_name]

    if not matches:
        raise RuntimeError(f"No existing report result found: {report_name}")

    matches = sorted(matches, key=lambda x: int(x["id"]), reverse=True)
    return matches

# ==============================
# REPORT DOWNLOAD
# ==============================

def download_report_result(session, headers, report_id, report_name, ext):
    output_dir = get_output_dir()

    safe_name = "".join(c for c in report_name if c not in r'\/:*?"<>|').strip()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{safe_name}_{timestamp}.{ext}"
    path = os.path.join(output_dir, filename)

    write_log(f"Downloading report {report_id} as {filename}")

    url = f"{SECURITY_CENTER}/report/{report_id}/download"
    r = session.post(url, headers=headers, verify=False)

    content_type = r.headers.get("Content-Type", "")
    size = len(r.content)

    write_log(f"Download content-type: {content_type}")
    write_log(f"Download size: {size} bytes")

    # Stop if the server returned JSON instead of a real file
    if "application/json" in content_type:
        write_log(f"Download failed, server returned JSON: {r.text}")
        return None

    # Stop if the file is empty
    if size == 0:
        write_log(f"Download failed, empty file returned for report {report_id}")
        return None

    with open(path, "wb") as f:
        f.write(r.content)

    write_log(f"Saved: {path}")
    return path

# ==============================
# DOWNLOAD LATEST AVAILABLE REPORT
# ==============================

def download_latest_available_report(session, headers, report_name, ext):
    matches = get_existing_reports_by_name(session, headers, report_name)

    for rpt in matches:
        report_id = str(rpt["id"])
        write_log(f"Trying report ID {report_id} for '{report_name}'")

        saved_path = download_report_result(session, headers, report_id, report_name, ext)
        if saved_path:
            write_log(f"Downloaded '{report_name}' successfully using report ID {report_id}")
            return saved_path

    raise RuntimeError(f"No downloadable report file found for: {report_name}")

# ==============================
# CSV TO XLSX CONVERSION
# ==============================

def csv_to_xlsx(csv_path):
    if not os.path.exists(csv_path):
        raise FileNotFoundError(f"CSV file not found: {csv_path}")

    xlsx_path = os.path.splitext(csv_path)[0] + ".xlsx"

    # Increase CSV field size limit for very large cells
    max_limit = sys.maxsize
    while True:
        try:
            csv.field_size_limit(max_limit)
            break
        except OverflowError:
            max_limit = int(max_limit / 10)

    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    with open(csv_path, newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        for row in reader:
            ws.append(row)

    # Apply header styling
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    header_font = Font(bold=True)

    if ws.max_row >= 1:
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

    # Freeze the first row
    ws.freeze_panes = "A2"

    # Apply auto filter
    ws.auto_filter.ref = ws.dimensions

    # Auto-fit column widths
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)

        for cell in col:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_length:
                max_length = len(value)

        ws.column_dimensions[col_letter].width = min(max_length + 2, 60)

    wb.save(xlsx_path)
    write_log(f"Converted CSV to Excel: {xlsx_path}")
    return xlsx_path

# ==============================
# OPTIONAL DEBUG HELPER
# ==============================

def list_existing_reports(session, headers):
    write_log("Listing all existing report results...")

    url = f"{SECURITY_CENTER}/report"
    r = session.get(url, headers=headers, verify=False)
    data = r.json()

    if data.get("error_code") != 0:
        raise RuntimeError(f"Failed to retrieve reports: {data}")

    reports = data["response"].get("usable", [])

    print("\n=== EXISTING REPORT RESULTS ===")
    for rpt in sorted(reports, key=lambda x: int(x.get("id", 0)), reverse=True):
        print(f"ID={rpt.get('id')} | NAME={rpt.get('name')}")
    print("================================\n")

# ==============================
# MAIN
# ==============================

def main():
    write_log("Starting ACAS automation")
    os.makedirs(EXPORT_DIR, exist_ok=True)

    session = requests.Session()
    token = get_token(session)
    headers = {"X-SecurityCenter": str(token)}

    print("TOKEN:", token)

    # Optional logging of the latest completed scan
    try:
        latest_scan = get_latest_completed_scan(session, headers)
        write_log(f"Using latest completed scan: {latest_scan['name']} ({latest_scan['id']})")
    except Exception as e:
        write_log(f"Scan lookup warning: {e}")

    # Uncomment this line if you want to print all existing report results for debugging
    # list_existing_reports(session, headers)

    # Download all existing reports from the latest available downloadable result
    for report in EXISTING_REPORTS:
        report_name = report["name"]
        ext = report["ext"]

        try:
            saved_path = download_latest_available_report(session, headers, report_name, ext)

            if not saved_path:
                write_log(f"FAILED: Existing report download -> {report_name} -> empty/error response")
            else:
                if ext.lower() == "csv":
                    try:
                        csv_to_xlsx(saved_path)
                    except Exception as convert_error:
                        write_log(f"FAILED: CSV to Excel conversion -> {report_name} -> {convert_error}")

        except Exception as e:
            write_log(f"FAILED: Existing report download -> {report_name} -> {e}")

    write_log("ACAS automation completed")


print("BEFORE MAIN")

if __name__ == "__main__":
    print("CALLING MAIN")
    main()

print("FILE END")